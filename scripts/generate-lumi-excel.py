# -*- coding: utf-8 -*-
"""
生成 D:/Lumi信息汇总.xlsx
- 数据源：public/data/internal/（对内版）
- 类型顺序：普通 → 异色卡 → 次元卡 → 王 → 幻境卡
- 内部排序：PokedexId 升序
- 进化链合并到一行（含分支进化）
- 立绘：优先 D:/噜咪立绘/（新高清图，命名多样），fallback public/images/avatars/CA_<Id>.png
- 三视图：D:/噜咪模型/Char_<Model>.(png|jpg)
"""
import json
import os
import sys
import io
import shutil
import re

# 强制 stdout 用 utf-8，避免 Windows GBK 撞车
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from PIL import Image as PILImage
import tempfile

# ---- 路径 ----
DATA_DIR = 'D:/lumiwiki/public/data/internal'
AVATAR_DIR = 'D:/lumiwiki/public/images/avatars'
NEW_AVATAR_DIR = 'D:/噜咪立绘'   # 新高清立绘目录（优先），命名多样：既可能是 6-7 位 Lumi ID 前缀，也可能是纯中文名
MODEL_DIR = 'D:/噜咪模型'
OUTPUT = 'D:/Lumi信息汇总.xlsx'

IMG_SIZE = 100                  # 立绘显示尺寸（像素），原图 256×256 无损嵌入
ROW_HEIGHT_PT_BASE = 80         # 基础行高（磅），三视图更高时按需拉伸
COL_WIDTH_IMG = 15              # 立绘列宽（Excel 单位，约对应 100px）
COL_WIDTH_TEXT = 16             # 文字列宽
COL_WIDTH_TYPE = 10             # 类型/赛季/属性列宽

# 三视图使用原图（无损嵌入），仅在 Excel 里控制显示宽度、按原比例算高度
MODEL_DISPLAY_WIDTH = 500       # 三视图显示宽度（像素）
COL_WIDTH_MODEL = 72            # 三视图列宽 (~500px)
PX_TO_PT = 0.75                 # 像素 → 磅 换算（Excel 行高单位是磅）

# 每段进化占用列数：中文名 / 英文名 / 属性 / 立绘 / 三视图
COLS_PER_STAGE = 5

TYPE_ORDER = {0: 0, 50: 1, 98: 2, 80: 3, 99: 4}
TYPE_LABEL = {0: '普通', 50: '异色卡', 98: '次元卡', 80: '王', 99: '幻境卡'}
SEASON_LABEL = {0: '未投放', 1: '主线', 2: 'S1', 3: 'S2', 4: 'S3', 5: 'S4'}

# LumiType（属性）映射，见 CLAUDE.md「枚举映射」章节
LUMI_TYPE = {
    0: '',
    1: '无', 2: '水', 3: '火', 4: '草', 5: '电', 6: '地',
    7: '飞', 8: '冰', 9: '龙', 10: '光', 11: '暗', 12: '格斗',
    13: '超能', 14: '妖精', 15: '钢', 16: '王', 17: '神',
}


def lumi_type_text(l):
    """返回噜咪属性文本：单属性 '水'，双属性 '水/火'。"""
    t1 = LUMI_TYPE.get(l.get('Type1', 0), '')
    t2 = LUMI_TYPE.get(l.get('Type2', 0), '')
    if t1 and t2:
        return f'{t1}/{t2}'
    return t1 or t2 or ''


def load_json(path):
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def build_model_index(model_dir):
    """一次性扫描 D:/噜咪模型/，返回 { normalized_name: filepath } 用于查找。"""
    idx = {}
    for f in os.listdir(model_dir):
        base, ext = os.path.splitext(f)
        if ext.lower() not in ('.png', '.jpg', '.jpeg'):
            continue
        # 去掉 Char_ 前缀
        key = base
        if key.startswith('Char_'):
            key = key[5:]
        idx[key.lower()] = os.path.join(model_dir, f)
    return idx


def find_model_file(model_name, model_idx):
    """通过 Model 字段查找三视图，尝试多种降级匹配。"""
    if not model_name:
        return None
    key = model_name.lower()
    # 1. 精确匹配
    if key in model_idx:
        return model_idx[key]
    # 2. 尝试去掉常见后缀
    for suffix in ('_01', '_3d', '_full', '_baby'):
        if key.endswith(suffix):
            base = key[:-len(suffix)]
            if base in model_idx:
                return model_idx[base]
    # 3. 忽略下划线的近似匹配
    key_flat = key.replace('_', '')
    for k, v in model_idx.items():
        if k.replace('_', '') == key_flat:
            return v
    # 4. 手工特例（数据表里模型名跟资源不完全一致）
    manual = {
        'chicken_gugu': 'chichen_gugu',
        'otter_steelhelm': 'ottersteelhelm',
        'deer_solaris': 'deer_solarise',
        'butterfly_meteor_01': 'butterfly_meteor',
        'wolf_frostwild_3d': 'wolf_frostwild',
        'deer_rainbow_full': 'deer_rainbow_full_d',
    }
    if key in manual and manual[key] in model_idx:
        return model_idx[manual[key]]
    # 5. 前缀模糊匹配
    for k, v in model_idx.items():
        if k.startswith(key) or key.startswith(k):
            return v
    return None


# ---- 新高清立绘匹配 ----
CJK_RE = re.compile(r'[一-鿿]+')
LEADING_DIGITS_RE = re.compile(r'^(\d{6,7})')
NOISE_WORDS = ['拷贝', '修改稿', '修改', '特效', '立绘']  # 都当作变体/副本标记


def clean_filename(base):
    """从文件名 base（无扩展名）提取纯中文噜咪名。"""
    s = base
    # 去掉括号内容 (1) （1） 等
    s = re.sub(r'[\(（].*?[\)）]', '', s)
    # 去掉噪声词
    for w in NOISE_WORDS:
        s = s.replace(w, '')
    # 去掉所有数字（前缀/中间/尾部）
    s = re.sub(r'\d+', '', s)
    # 拼接所有 CJK 段（丢掉英文字母/空格/下划线/连字符/·等）
    return ''.join(CJK_RE.findall(s))


def try_match_name(cjk, name_to_ids):
    """按中文名找 Lumi ID list，处理 `异色X`/`X异色` → `X·异` 变体。"""
    if not cjk:
        return []
    if cjk in name_to_ids:
        return name_to_ids[cjk]
    # 异色变体：文件名写 `异色金刚狼` 或 `金刚狼异色`，游戏数据里是 `金刚狼·异`
    for pfx in ('异色',):
        if cjk.startswith(pfx):
            core = cjk[len(pfx):]
            key = f'{core}·异'
            if key in name_to_ids:
                return name_to_ids[key]
        if cjk.endswith(pfx):
            core = cjk[:-len(pfx)]
            key = f'{core}·异'
            if key in name_to_ids:
                return name_to_ids[key]
    return []


def build_new_avatar_index(new_dir, zh, lumi_by_id):
    """
    扫描新高清立绘目录，返回 ({lumi_id: filepath}, unmatched_files)。
    命中优先级：
      1. 文件名以 6-7 位数字开头且该数字是有效 Lumi ID → 直接用
      2. 提取中文名 → zh-CN 反查（含 异色 变体处理）
    同一个 Lumi 有多个候选文件时，选 priority 最小、无噪声后缀（拷贝/修改/带尾部数字）的。
    """
    if not os.path.isdir(new_dir):
        print(f'⚠️  新立绘目录不存在: {new_dir}，跳过。')
        return {}, []

    # 建 zh-CN 中文名 → [lumi_id] 反向索引
    name_to_ids = {}
    for l in lumi_by_id.values():
        zh_name = zh.get(l.get('Name', ''))
        if zh_name:
            name_to_ids.setdefault(zh_name, []).append(l['Id'])

    # candidates[lid] = [(priority, has_noise, filepath), ...]
    candidates = {}
    unmatched = []

    for fname in sorted(os.listdir(new_dir)):
        if not fname.lower().endswith(('.png', '.jpg', '.jpeg')):
            continue
        base = os.path.splitext(fname)[0]
        full = os.path.join(new_dir, fname)

        has_noise_word = any(w in base for w in NOISE_WORDS) or '(' in base or '（' in base
        # 去括号后判断末尾是否还带数字（`蜜桃1`、`猫鲨11`）
        base_no_paren = re.sub(r'[\(（].*?[\)）]', '', base).strip()
        has_trailing_num = bool(re.search(r'\d+$', base_no_paren))

        matched_ids = None
        priority = None

        # 尝试 1：6-7 位数字前缀（7 位不命中时降级到 6 位再试）
        m = LEADING_DIGITS_RE.match(base)
        if m:
            digits = m.group(1)
            lid = int(digits)
            if lid in lumi_by_id:
                matched_ids = [lid]
                priority = 0
            elif len(digits) == 7:
                lid6 = int(digits[:6])
                if lid6 in lumi_by_id:
                    matched_ids = [lid6]
                    priority = 0

        # 尝试 2：中文名匹配
        if matched_ids is None:
            cjk = clean_filename(base)
            ids = try_match_name(cjk, name_to_ids)
            if ids:
                matched_ids = ids
                priority = 1

        if matched_ids is None:
            unmatched.append(fname)
            continue

        # priority=0（前缀命中）时不考虑 has_trailing_num（前缀本身就是数字）
        noise_flag = has_noise_word if priority == 0 else (has_noise_word or has_trailing_num)
        for lid in matched_ids:
            candidates.setdefault(lid, []).append((priority, noise_flag, full))

    # 每个 Lumi 挑最佳文件：priority 升序 → has_noise 升序 → 文件名字典序
    result = {}
    for lid, lst in candidates.items():
        lst.sort(key=lambda x: (x[0], x[1], x[2]))
        result[lid] = lst[0][2]

    return result, unmatched


def find_avatar(lumi_id, new_idx):
    """只用新高清立绘目录；没有就置空（不 fallback 旧 CA_）。返回 (path, source)。"""
    p = new_idx.get(lumi_id)
    if p and os.path.exists(p):
        return p, 'new'
    return None, 'none'


def resize_image(src_path, size, tmp_dir):
    """缩放图片到正方形（保持比例，用透明背景填充）。返回临时文件路径。"""
    try:
        img = PILImage.open(src_path).convert('RGBA')
    except Exception as e:
        print(f'  ⚠️ 读取图片失败: {src_path} - {e}')
        return None
    # 等比缩放
    img.thumbnail((size, size), PILImage.LANCZOS)
    # 贴到正方形透明画布
    canvas = PILImage.new('RGBA', (size, size), (255, 255, 255, 0))
    off_x = (size - img.width) // 2
    off_y = (size - img.height) // 2
    canvas.paste(img, (off_x, off_y), img)
    # 存到临时文件
    tmp_path = os.path.join(tmp_dir, f'tmp_{os.path.basename(src_path)}.png')
    canvas.save(tmp_path, 'PNG')
    return tmp_path


def get_evo_targets(lumi_id, evo_data):
    """返回该 Lumi 的直接进化目标列表（去重），以及是否为分支。"""
    targets = []
    seen = set()
    for e in evo_data:
        if e['Lumi'] != lumi_id:
            continue
        tid = e.get('evoLumiID', 0)
        if tid > 0 and tid not in seen:
            targets.append(tid)
            seen.add(tid)
        for _, gid in e.get('GenderEvo', []):
            if gid not in seen:
                targets.append(gid)
                seen.add(gid)
    is_branch = len(targets) > 1
    return targets, is_branch


def build_chain(start_id, evo_data, visited=None):
    """
    返回该起始 Lumi 的进化链（Id 列表）。
    - 线性进化：[base, evo1, evo2, ...]
    - 分支进化：[base, branch1, branch2, ...]（分支占用连续槽位）
    """
    if visited is None:
        visited = set()
    chain = [start_id]
    if start_id in visited:
        return chain
    visited.add(start_id)
    targets, is_branch = get_evo_targets(start_id, evo_data)
    if not targets:
        return chain
    if is_branch:
        # 所有分支平铺（分支目标经确认均为终点，不再深挖）
        for t in targets:
            chain.append(t)
            visited.add(t)
    else:
        sub = build_chain(targets[0], evo_data, visited)
        chain.extend(sub)
    return chain


def main():
    print('📖 读取数据...')
    lumi_data = load_json(f'{DATA_DIR}/Lumi.json')
    evo_data = load_json(f'{DATA_DIR}/LumiEvolution.json')
    zh = load_json(f'{DATA_DIR}/zh-CN.json')
    en = load_json(f'{DATA_DIR}/en.json')
    lumi_by_id = {l['Id']: l for l in lumi_data}

    print(f'   噜咪总数 {len(lumi_data)}，进化记录 {len(evo_data)}')

    # 建立三视图索引
    print('🖼️  扫描三视图目录...')
    model_idx = build_model_index(MODEL_DIR)
    print(f'   索引文件 {len(model_idx)} 个')

    # 建立新高清立绘索引（优先来源）
    print('✨ 扫描新高清立绘目录...')
    new_avatar_idx, unmatched_new = build_new_avatar_index(NEW_AVATAR_DIR, zh, lumi_by_id)
    print(f'   新目录匹配到 {len(new_avatar_idx)} 只 Lumi，未匹配 {len(unmatched_new)} 个文件')

    # 找出所有被进化到的 Id（不作为首形态出现）
    evolved_into = set()
    for e in evo_data:
        if e.get('evoLumiID', 0) > 0:
            evolved_into.add(e['evoLumiID'])
        for _, lid in e.get('GenderEvo', []):
            evolved_into.add(lid)

    # ---- 组装行 ----
    rows = []

    # 普通（CardBack=0）：首形态 + 进化链
    normal_first = [l for l in lumi_data
                    if l.get('CardBack', 0) == 0
                    and l.get('LumiTag', 0) != 0
                    and l['Id'] not in evolved_into]
    for l in normal_first:
        chain = build_chain(l['Id'], evo_data)
        rows.append({
            'type': 0,
            'season': l.get('LumiTag', 0),
            'pokedex': l.get('PokedexId', 0),
            'chain': chain,
        })

    # 特殊卡（异色/3D/王/幻境）：独立一行
    for cb in (50, 98, 80, 99):
        specials = [l for l in lumi_data
                    if l.get('CardBack', 0) == cb
                    and l.get('LumiTag', 0) != 0]
        for l in specials:
            rows.append({
                'type': cb,
                'season': l.get('LumiTag', 0),
                'pokedex': l.get('PokedexId', 0),
                'chain': [l['Id']],
            })

    # 排序：类型 → PokedexId
    rows.sort(key=lambda r: (TYPE_ORDER[r['type']], r['pokedex']))

    max_chain_len = max(len(r['chain']) for r in rows)
    print(f'📊 总行数 {len(rows)}，最长进化链 {max_chain_len} 段')

    # ---- 生成 Excel ----
    print('📝 生成 Excel...')
    wb = Workbook()
    ws = wb.active
    ws.title = '噜咪信息汇总'

    # 表头
    header = ['类型', '获取途径']
    for i in range(1, max_chain_len + 1):
        stage = '第一形态' if i == 1 else f'第{["一","二","三","四","五","六","七","八","九"][i-1]}进化'
        header += [f'{stage}中文名', f'{stage}英文名', f'{stage}属性', f'{stage}立绘', f'{stage}三视图']
    for i, h in enumerate(header, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = PatternFill('solid', fgColor='FFE699')

    # 列宽
    ws.column_dimensions['A'].width = COL_WIDTH_TYPE
    ws.column_dimensions['B'].width = COL_WIDTH_TYPE
    for i in range(max_chain_len):
        base_col = 3 + i * COLS_PER_STAGE
        ws.column_dimensions[get_column_letter(base_col)].width = COL_WIDTH_TEXT       # 中文
        ws.column_dimensions[get_column_letter(base_col + 1)].width = COL_WIDTH_TEXT   # 英文
        ws.column_dimensions[get_column_letter(base_col + 2)].width = COL_WIDTH_TYPE   # 属性
        ws.column_dimensions[get_column_letter(base_col + 3)].width = COL_WIDTH_IMG    # 立绘 (~100px 显示，原图存储)
        ws.column_dimensions[get_column_letter(base_col + 4)].width = COL_WIDTH_MODEL  # 三视图 (~500px 显示，原图存储)

    ws.row_dimensions[1].height = 24

    # 用 TemporaryDirectory 保存缩放后的图片，写完统一清理
    tmp_dir = tempfile.mkdtemp(prefix='lumi_xlsx_')
    missing_avatars = []
    missing_models = []
    stats_new = 0    # 用了新高清立绘的 Lumi 数
    stats_old = 0    # fallback 到旧 CA_ 的 Lumi 数
    fallback_lids = []  # 具体是哪些 Lumi fallback 了

    total_rows = len(rows)
    for row_idx, r in enumerate(rows, start=2):
        if (row_idx - 1) % 20 == 0 or row_idx == total_rows + 1:
            print(f'  处理行 {row_idx - 1}/{total_rows}...')
        ws.cell(row=row_idx, column=1, value=TYPE_LABEL.get(r['type'], '?')).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_idx, column=2, value=SEASON_LABEL.get(r['season'], '?')).alignment = Alignment(horizontal='center', vertical='center')

        # 收集本行三视图最大高度，用于设置行高
        row_max_model_h_px = 0

        for stage_i, lid in enumerate(r['chain']):
            base_col = 3 + stage_i * COLS_PER_STAGE
            l = lumi_by_id.get(lid)
            if not l:
                continue
            name_key = l.get('Name', '')
            zh_name = zh.get(name_key, name_key)
            en_name = en.get(name_key, name_key)
            attr_text = lumi_type_text(l)
            ws.cell(row=row_idx, column=base_col, value=zh_name).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row=row_idx, column=base_col + 1, value=en_name).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row=row_idx, column=base_col + 2, value=attr_text).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # 立绘：优先新高清（原图无损嵌入），fallback 旧 CA_，仅显示 100×100
            avatar_path, avatar_src = find_avatar(lid, new_avatar_idx)
            if avatar_path:
                try:
                    img = XLImage(avatar_path)  # openpyxl 直接嵌原文件字节
                    img.width = IMG_SIZE
                    img.height = IMG_SIZE
                    anchor_cell = f'{get_column_letter(base_col + 3)}{row_idx}'
                    ws.add_image(img, anchor_cell)
                    if avatar_src == 'new':
                        stats_new += 1
                    else:
                        stats_old += 1
                        fallback_lids.append((lid, zh_name))
                except Exception as e:
                    print(f'  ⚠️ 立绘嵌入失败: {avatar_path} - {e}')
                    missing_avatars.append(lid)
            else:
                missing_avatars.append(lid)

            # 三视图：插入原图字节，不做任何缩放；仅通过 img.width/height 控制 Excel 里的显示尺寸
            model_name = l.get('Model', '')
            model_path = find_model_file(model_name, model_idx)
            if model_path:
                try:
                    with PILImage.open(model_path) as pim:
                        orig_w, orig_h = pim.size
                    img = XLImage(model_path)
                    display_h = int(MODEL_DISPLAY_WIDTH * orig_h / orig_w)
                    img.width = MODEL_DISPLAY_WIDTH
                    img.height = display_h
                    anchor_cell = f'{get_column_letter(base_col + 4)}{row_idx}'
                    ws.add_image(img, anchor_cell)
                    row_max_model_h_px = max(row_max_model_h_px, display_h)
                except Exception as e:
                    print(f'  ⚠️ 三视图嵌入失败: {model_path} - {e}')
                    missing_models.append((lid, model_name))
            else:
                missing_models.append((lid, model_name))

        # 行高：max(基础, 本行三视图最高像素 × 0.75 → pt)
        row_height_pt = max(ROW_HEIGHT_PT_BASE, row_max_model_h_px * PX_TO_PT)
        ws.row_dimensions[row_idx].height = row_height_pt

    # 冻结表头
    ws.freeze_panes = 'C2'

    print(f'💾 保存到 {OUTPUT}')
    wb.save(OUTPUT)

    # 清理临时缩放图片
    shutil.rmtree(tmp_dir, ignore_errors=True)

    # 报告
    print('')
    print(f'✅ 完成！共 {len(rows)} 行')
    print(f'🖼️  立绘来源：新高清 {stats_new} · 旧 CA_ fallback {stats_old} · 缺失 {len(missing_avatars)}')
    if fallback_lids:
        print(f'   fallback 到旧图的 Lumi:')
        for lid, name in fallback_lids:
            print(f'    - {lid} {name}')
    if missing_avatars:
        print(f'⚠️  缺少立绘 {len(missing_avatars)} 个: {missing_avatars[:10]}{"..." if len(missing_avatars)>10 else ""}')
    if missing_models:
        print(f'⚠️  缺少三视图 {len(missing_models)} 个:')
        for lid, m in missing_models:
            print(f'    Id={lid} Model={m}')
    if unmatched_new:
        print(f'')
        print(f'ℹ️  新目录里 {len(unmatched_new)} 个文件没匹配到任何 Lumi（噪声/变体/未启用，均忽略）:')
        for f in unmatched_new[:20]:
            print(f'    - {f}')
        if len(unmatched_new) > 20:
            print(f'    ... 还有 {len(unmatched_new) - 20} 个')


if __name__ == '__main__':
    main()

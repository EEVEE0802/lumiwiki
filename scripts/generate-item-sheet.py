# -*- coding: utf-8 -*-
"""
更新 D:/Lumi信息汇总.xlsx 的「道具信息」sheet
- 数据源：public/data/internal/Item.json（对内版）
- 范围：isShow=True 的物品（游戏内图鉴显示的 147 个）
- 排序：type → order → key1
- 图标：D:/lumiwiki/public/images/internal/items/<icon>.png
- ⚠️ 只操作 sheet 2，sheet 1（噜咪信息汇总）保持不变
"""
import json
import os
import sys
import io
import shutil
import tempfile

if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image as PILImage

DATA_DIR = 'D:/lumiwiki/public/data/internal'
ITEM_ICON_DIR = 'D:/lumiwiki/public/images/internal/items'
OUTPUT = 'D:/Lumi信息汇总.xlsx'
SHEET_NAME = '道具信息'

IMG_SIZE = 100
ROW_HEIGHT_PT = 80
COL_NAME_WIDTH = 22
COL_DES_WIDTH = 60
COL_IMG_WIDTH = 15


def load_json(path):
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def resize_image(src_path, size, tmp_dir):
    try:
        img = PILImage.open(src_path).convert('RGBA')
    except Exception as e:
        print(f'  ⚠️ 读图失败: {src_path} - {e}')
        return None
    img.thumbnail((size, size), PILImage.LANCZOS)
    canvas = PILImage.new('RGBA', (size, size), (255, 255, 255, 0))
    off_x = (size - img.width) // 2
    off_y = (size - img.height) // 2
    canvas.paste(img, (off_x, off_y), img)
    tmp_path = os.path.join(tmp_dir, f'tmp_{os.path.basename(src_path)}.png')
    canvas.save(tmp_path, 'PNG')
    return tmp_path


def find_icon(icon_name):
    if not icon_name:
        return None
    for ext in ('.png', '.jpg', '.jpeg'):
        p = os.path.join(ITEM_ICON_DIR, f'{icon_name}{ext}')
        if os.path.exists(p):
            return p
    return None


def main():
    print('📖 读取物品数据...')
    items = load_json(f'{DATA_DIR}/Item.json')
    zh = load_json(f'{DATA_DIR}/zh-CN.json')
    print(f'   物品总数 {len(items)}')

    # 过滤 & 排序
    show_items = [i for i in items if i.get('isShow', False)]
    show_items.sort(key=lambda i: (i.get('type', 0), i.get('order', 0), i.get('key1', 0)))
    print(f'   isShow=True: {len(show_items)} 个')

    # 加载现有 xlsx（保留 sheet 1）
    print(f'📂 加载 {OUTPUT} ...')
    wb = load_workbook(OUTPUT)
    print(f'   现有 sheet: {wb.sheetnames}')

    # 定位 sheet 2
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        # 清空数据行（保留表头行 1）
        max_r = ws.max_row
        max_c = ws.max_column
        if max_r > 1:
            ws.delete_rows(2, max_r - 1)
        # 清掉旧的图片
        ws._images = []
    else:
        ws = wb.create_sheet(SHEET_NAME)

    # 表头（若不存在则补，若已存在保持用户设置的）
    if not ws.cell(1, 1).value:
        headers = ['名字', '用途', '图片']
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.fill = PatternFill('solid', fgColor='FFE699')
        ws.row_dimensions[1].height = 24

    # 列宽
    ws.column_dimensions['A'].width = COL_NAME_WIDTH
    ws.column_dimensions['B'].width = COL_DES_WIDTH
    ws.column_dimensions['C'].width = COL_IMG_WIDTH

    # 填数据
    tmp_dir = tempfile.mkdtemp(prefix='lumi_item_')
    missing_icons = []
    missing_names = []
    missing_dess = []

    for row_idx, item in enumerate(show_items, start=2):
        ws.row_dimensions[row_idx].height = ROW_HEIGHT_PT
        name_key = item.get('name', '')
        des_key = item.get('des', '')
        zh_name = zh.get(name_key)
        zh_des = zh.get(des_key)
        icon = item.get('icon', '')

        display_name = zh_name if zh_name else f'[缺翻译] {name_key}'
        display_des = zh_des if zh_des else (f'[缺描述] {des_key}' if des_key else '')

        ws.cell(row=row_idx, column=1, value=display_name).alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=row_idx, column=2, value=display_des).alignment = Alignment(
            horizontal='left', vertical='center', wrap_text=True)

        if not zh_name:
            missing_names.append((item.get('key1'), name_key))
        if des_key and not zh_des:
            missing_dess.append((item.get('key1'), des_key))

        icon_path = find_icon(icon)
        if icon_path:
            resized = resize_image(icon_path, IMG_SIZE, tmp_dir)
            if resized:
                img = XLImage(resized)
                img.width = IMG_SIZE
                img.height = IMG_SIZE
                ws.add_image(img, f'C{row_idx}')
        else:
            missing_icons.append((item.get('key1'), icon, display_name))

    ws.freeze_panes = 'A2'

    print(f'💾 保存到 {OUTPUT}')
    wb.save(OUTPUT)

    shutil.rmtree(tmp_dir, ignore_errors=True)

    # 报告
    print('')
    print(f'✅ 完成！道具信息 sheet 写入 {len(show_items)} 行')
    if missing_names:
        print(f'⚠️  缺中文名 {len(missing_names)} 个:')
        for k, n in missing_names:
            print(f'    key1={k} name_key={n}')
    if missing_dess:
        print(f'⚠️  缺描述 {len(missing_dess)} 个:')
        for k, d in missing_dess:
            print(f'    key1={k} des_key={d}')
    if missing_icons:
        print(f'⚠️  缺图标 {len(missing_icons)} 个:')
        for k, i, n in missing_icons:
            print(f'    key1={k} icon={i} name={n}')


if __name__ == '__main__':
    main()
